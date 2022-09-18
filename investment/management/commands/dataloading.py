from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from investment.models import Account, Asset, AssetGroup
from django.contrib.auth import get_user_model


class Command(BaseCommand):
    help = 'AssetGroup Data Loading'

    def handle(self, *args, **options):
        '''
        엑셀 데이터 로드 및 DB 저장
        '''
        asset_group_wb = load_workbook('investment/data_files/asset_group_info_set.xlsx')
        asset_group_ws = asset_group_wb.active

        name_col = asset_group_ws['A']
        isin_col = asset_group_ws['B']
        group_col = asset_group_ws['C']

        name_list = []
        isin_list = []
        group_list = []

        for cell in name_col[1:]:
            name_list.append(cell.value)

        for cell in isin_col[1:]:
            isin_list.append(cell.value)

        for cell in group_col[1:]:
            group_list.append(cell.value)

        '''
        # 1. AseetGroup 테이블
            - 그룹 테이블 저장
            - asset_group_info_set.xlsx 데이터 저장
        '''
        for i in range(len(name_list)):
            assetGroup = AssetGroup.objects.filter(isin=isin_list[i])
            if assetGroup:
                continue
            else:
                AssetGroup(asset_name=name_list[i]
                           , isin=isin_list[i]
                           , group_name=group_list[i]).save()

        account_asset_wb = load_workbook('investment/data_files/account_asset_info_set.xlsx')
        account_asset_ws = account_asset_wb.active

        user_name_col = account_asset_ws['A']
        brokerage_col = account_asset_ws['B']
        asset_account_number_col = account_asset_ws['C']
        account_name_col = account_asset_ws['D']
        asset_isin_col = account_asset_ws['E']
        current_price_col = account_asset_ws['F']
        quantity_col = account_asset_ws['G']

        user_name_list = []
        brokerage_list = []
        asset_account_number_list = []
        account_name_list = []
        asset_isin_list = []
        current_price_list = []
        quantity_list = []

        for cell in user_name_col[1:]:
            user_name_list.append(cell.value)

        for cell in brokerage_col[1:]:
            brokerage_list.append(cell.value)

        for cell in asset_account_number_col[1:]:
            asset_account_number_list.append(cell.value)

        for cell in account_name_col[1:]:
            account_name_list.append(cell.value)

        for cell in asset_isin_col[1:]:
            asset_isin_list.append(cell.value)

        for cell in current_price_col[1:]:
            current_price_list.append(cell.value)

        for cell in quantity_col[1:]:
            quantity_list.append(cell.value)
        '''
        # 2. User 테이블
            - asset_asset_info_set.xlsx의 사용자 이름으로 저장
            - 비밀번호 0000으로 임시 저장 
        '''
        user_name_set = list(set(user_name_list))
        password = 'pbkdf2_sha256$260000$BAG0oHkWNYZqurszaCLvMT$zZjijq9PrNCJMac8kYHXJQN/6k0KndAeVO/UnIG20Xk='  # 0000

        # 사용자 이름을 username(unique) 값으로 사용하기 때문에 중복 제거
        for i in range(len(user_name_set)):
            user = get_user_model().objects.filter(username=user_name_set[i])
            if user:
                continue
            else:
                get_user_model()(username=user_name_set[i]
                                 , password=password).save()

        account_basic_wb = load_workbook('investment/data_files/account_basic_info_set.xlsx')
        account_basic_ws = account_basic_wb.active

        basic_account_number_col = account_basic_ws['A']
        total_amount_col = account_basic_ws['B']

        basic_account_number_list = []
        total_amount_list = []

        for cell in basic_account_number_col[1:]:
            basic_account_number_list.append(cell.value)

        for cell in total_amount_col[1:]:
            total_amount_list.append(cell.value)
        '''
        # 3. Account 테이블
            - get_user_model()에서 id 참조
            - asset_basic_info_set.xlsx에서 total_amount 참조
        '''
        for i in range(len(user_name_list)):
            user = get_user_model().objects.filter(username=user_name_list[i])
            total_amount = 0

            for j in range(len(basic_account_number_list)):
                if basic_account_number_list[j] == asset_account_number_list[i]:
                    total_amount = total_amount_list[j]

            account = Account.objects.filter(account_number=asset_account_number_list[i])
            if account.exists():
                continue
            else:
                Account(account_number=asset_account_number_list[i]
                        , account_name=account_name_list[i]
                        , brokerage=brokerage_list[i]
                        , total_amount=total_amount
                        , user=user[0]).save()
        '''
        # 4. Asset 테이블
            - Account에서 id 참조 
            - AssetGroup에서 id 참조
        '''
        for i in range(len(asset_account_number_list)):
            account = Account.objects.get(account_number=asset_account_number_list[i])
            assetGroup = AssetGroup.objects.get(isin=asset_isin_list[i])

            asset = Asset.objects.filter(account=account, group=assetGroup)
            if asset.exists():
                continue
            else:
                Asset(account=account
                      , group=assetGroup
                      , current_price=current_price_list[i]
                      , quantity=quantity_list[i]).save()
